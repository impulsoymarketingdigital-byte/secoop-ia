#!/usr/bin/env python3
"""
JIREHAI - SECOP AI Analyst
exporters.py — Genera exportaciones del análisis en múltiples formatos

Genera:
  - analysis.json          — Resultado completo estructurado
  - informe_completo.xlsx  — Informe único con TODAS las hojas (Resumen, Cronograma,
                             Requisitos, Financiero, Productos, Riesgos, Documentos)
  - informe_completo.pdf   — PDF unificado (via reportlab o fpdf2, con fallback HTML)
  - executive_summary.md   — Resumen ejecutivo Markdown
  - legal_analysis.md      — Análisis jurídico detallado
  - risk_report.md         — Reporte de riesgos
  - product_summary.json   — Productos y especificaciones
"""
import json
import logging
import os
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)
DATE_FMT = "%Y-%m-%d %H:%M"


# ─────────────────────────────────────────────────────────────────────────────
# Punto de entrada
# ─────────────────────────────────────────────────────────────────────────────

def export_all(analysis_data: dict, output_dir: Path) -> dict:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    generated = {}

    # 1. JSON completo
    try:
        p = output_dir / "analysis.json"
        _write_json(analysis_data, p)
        generated["json"] = str(p)
        logger.info("✅ analysis.json generado")
    except Exception as e:
        logger.error(f"Error analysis.json: {e}")

    # 2. Excel unificado (informe completo)
    try:
        p = output_dir / "informe_completo.xlsx"
        _build_full_xlsx(analysis_data, p)
        generated["compliance_matrix"] = str(p)   # clave legacy para compatibilidad
        generated["informe_xlsx"] = str(p)
        logger.info("✅ informe_completo.xlsx generado")
    except Exception as e:
        logger.warning(f"No se pudo generar informe_completo.xlsx: {e}")
        # fallback: solo compliance básico
        try:
            p = output_dir / "compliance_matrix.xlsx"
            _build_compliance_xlsx_basic(analysis_data, p)
            generated["compliance_matrix"] = str(p)
            logger.info("✅ compliance_matrix.xlsx generado (fallback)")
        except Exception as e2:
            logger.warning(f"Tampoco se pudo generar compliance básico: {e2}")

    # 3. PDF unificado
    try:
        p = output_dir / "informe_completo.pdf"
        _build_full_pdf(analysis_data, p)
        generated["informe_pdf"] = str(p)
        logger.info("✅ informe_completo.pdf generado")
    except Exception as e:
        logger.warning(f"PDF no generado: {e}")

    # 4. Resumen ejecutivo MD
    try:
        p = output_dir / "executive_summary.md"
        p.write_text(_build_executive_summary(analysis_data), encoding="utf-8")
        generated["executive_summary"] = str(p)
        logger.info("✅ executive_summary.md generado")
    except Exception as e:
        logger.error(f"Error executive_summary.md: {e}")

    # 5. Análisis jurídico MD
    try:
        p = output_dir / "legal_analysis.md"
        p.write_text(_build_legal_md(analysis_data.get("legal", {}),
                                     analysis_data.get("process_info", {})), encoding="utf-8")
        generated["legal_analysis"] = str(p)
        logger.info("✅ legal_analysis.md generado")
    except Exception as e:
        logger.error(f"Error legal_analysis.md: {e}")

    # 6. Reporte de riesgos MD
    try:
        p = output_dir / "risk_report.md"
        p.write_text(_build_risk_report_md(analysis_data.get("risks", {}),
                                           analysis_data.get("process_info", {})), encoding="utf-8")
        generated["risk_report"] = str(p)
        logger.info("✅ risk_report.md generado")
    except Exception as e:
        logger.error(f"Error risk_report.md: {e}")

    # 7. Productos JSON
    try:
        p = output_dir / "product_summary.json"
        _write_json(analysis_data.get("products", {}), p)
        generated["product_summary"] = str(p)
        logger.info("✅ product_summary.json generado")
    except Exception as e:
        logger.error(f"Error product_summary.json: {e}")

    return generated


# ─────────────────────────────────────────────────────────────────────────────
# Excel: informe completo multi-hoja
# ─────────────────────────────────────────────────────────────────────────────

def _build_full_xlsx(data: dict, output_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_DATE_DATETIME

    wb = Workbook()
    wb.remove(wb.active)

    process  = data.get("process_info", {})
    legal    = data.get("legal", {})
    financial= data.get("financial", {})
    risks    = data.get("risks", {})
    products = data.get("products", {})
    docs     = data.get("documents", [])

    # ── Paleta de colores ────────────────────────────────────────────────
    C_BLUE_DARK   = "1F4E79"
    C_BLUE_MED    = "2E75B6"
    C_BLUE_LIGHT  = "BDD7EE"
    C_GREEN_DARK  = "375623"
    C_GREEN_MED   = "70AD47"
    C_GREEN_LIGHT = "E2EFDA"
    C_RED_DARK    = "833C00"
    C_RED_MED     = "C00000"
    C_RED_LIGHT   = "FCE4D6"
    C_ORANGE      = "ED7D31"
    C_YELLOW      = "FFF2CC"
    C_GREY_LIGHT  = "F2F2F2"
    C_WHITE       = "FFFFFF"

    thin = Border(
        left=Side(style="thin"),   right=Side(style="thin"),
        top=Side(style="thin"),    bottom=Side(style="thin"),
    )
    thick_bottom = Border(bottom=Side(style="medium"))

    def hdr(cell, bg=C_BLUE_DARK, fg="FFFFFF", bold=True, size=10):
        cell.font      = Font(bold=bold, color=fg, size=size)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin

    def cell_style(cell, bold=False, bg=None, align="left", wrap=True):
        cell.font      = Font(bold=bold, size=9)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border    = thin
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    def title_row(ws, text, row, cols, bg=C_BLUE_DARK):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin

    def set_col_widths(ws, widths):
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 1 — Resumen Ejecutivo
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("1. Resumen Ejecutivo")
    set_col_widths(ws1, [30, 70])

    title_row(ws1, "📋 ANÁLISIS SECOP II — INFORME COMPLETO", 1, 2)

    filas_resumen = [
        ("INFORMACIÓN DEL PROCESO", "", C_BLUE_MED),
        ("Referencia del proceso",  process.get("referencia", "N/A"), None),
        ("Entidad contratante",     process.get("entidad", "N/A"), None),
        ("Objeto del contrato",     legal.get("objeto_contrato", "N/A"), None),
        ("Alcance",                 legal.get("alcance", "Ver pliego"), None),
        ("Modalidad de selección",  legal.get("modalidad_seleccion", "N/A"), None),
        ("Régimen jurídico",        legal.get("regimen_juridico", "N/A"), None),
        ("Valor estimado",          legal.get("valor_estimado", "N/A"), None),
        ("Duración del contrato",   legal.get("duracion_contrato", "N/A"), None),
        ("Lugar de ejecución",      legal.get("lugar_ejecucion", "N/A"), None),
        ("Tipo de bien/servicio",   products.get("tipo_bien_servicio", "N/A"), None),
        ("RESULTADO DEL ANÁLISIS IA", "", C_GREEN_DARK),
        ("Nivel de riesgo global",  risks.get("nivel_riesgo_global", "N/A"), None),
        ("Score de riesgo",         f"{risks.get('score_riesgo', 0)}/100", None),
        ("Documentos descargados",  str(process.get("documentos_descargados", 0)), None),
        ("Documentos analizados",   str(process.get("documentos_parseados", 0)), None),
        ("Texto extraído (chars)",  str(process.get("texto_extraido_chars", 0)), None),
        ("Fecha del análisis",      datetime.now().strftime(DATE_FMT), None),
    ]

    for r_idx, (k, v, sec_bg) in enumerate(filas_resumen, 2):
        ca = ws1.cell(row=r_idx, column=1, value=k)
        cb = ws1.cell(row=r_idx, column=2, value=v)
        if sec_bg:
            hdr(ca, bg=sec_bg, size=9)
            hdr(cb, bg=sec_bg, size=9)
        else:
            cell_style(ca, bold=True, bg=C_GREY_LIGHT)
            cell_style(cb)

    ws1.row_dimensions[1].height = 25
    for r in range(2, len(filas_resumen) + 2):
        ws1.row_dimensions[r].height = 18

    # ── Resumen ejecutivo de riesgos ──────────────────────────────────
    rs_text = risks.get("resumen_ejecutivo_riesgos", "")
    if rs_text:
        r_start = len(filas_resumen) + 3
        ws1.merge_cells(start_row=r_start, start_column=1, end_row=r_start, end_column=2)
        c = ws1.cell(row=r_start, column=1, value="RESUMEN EJECUTIVO DE RIESGOS")
        hdr(c, bg=C_RED_DARK, size=9)
        ws1.cell(row=r_start, column=2)

        r_start += 1
        ws1.merge_cells(start_row=r_start, start_column=1, end_row=r_start+3, end_column=2)
        c = ws1.cell(row=r_start, column=1, value=rs_text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = thin
        ws1.row_dimensions[r_start].height = 60

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 2 — Cronograma
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("2. Cronograma")
    set_col_widths(ws2, [40, 22, 22, 40])
    title_row(ws2, "📅 CRONOGRAMA DEL PROCESO", 1, 4, bg=C_BLUE_MED)

    hdrs2 = ["Evento / Actividad", "Fecha inicio", "Fecha límite", "Observaciones"]
    for ci, h in enumerate(hdrs2, 1):
        hdr(ws2.cell(row=2, column=ci, value=h), bg=C_BLUE_MED)

    cronograma = legal.get("cronograma", [])
    for ri, ev in enumerate(cronograma, 3):
        vals = [
            ev.get("evento", ""),
            ev.get("fecha", ev.get("fecha_inicio", "")),
            ev.get("fecha_limite", ev.get("fecha", "")),
            ev.get("observaciones", ""),
        ]
        bg = C_YELLOW if ri % 2 == 0 else C_WHITE
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            cell_style(c, bg=bg)

    if not cronograma:
        ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
        c = ws2.cell(row=3, column=1, value="No se identificaron fechas en los documentos analizados")
        cell_style(c, bg=C_YELLOW)

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 3 — Requisitos Habilitantes
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("3. Requisitos Habilitantes")
    set_col_widths(ws3, [15, 45, 45, 20, 12])
    title_row(ws3, "🔑 REQUISITOS HABILITANTES", 1, 5, bg=C_GREEN_DARK)

    hdrs3 = ["Tipo", "Requisito", "Cómo acreditar / Documento exigido", "Valor mínimo", "Obligatorio"]
    for ci, h in enumerate(hdrs3, 1):
        hdr(ws3.cell(row=2, column=ci, value=h), bg=C_GREEN_MED)

    req_hab = legal.get("requisitos_habilitantes", {})
    row = 3
    for tipo, reqs, bg in [
        ("Jurídico",      req_hab.get("juridicos", []),      C_GREEN_LIGHT),
        ("Técnico",       req_hab.get("tecnicos", []),       C_YELLOW),
        ("Organizacional",req_hab.get("organizacionales", []),C_GREY_LIGHT),
    ]:
        for r in reqs:
            vals = [
                tipo,
                r.get("requisito", ""),
                r.get("documento_exigido", "") or r.get("como_acreditar", ""),
                "",
                "Sí" if r.get("obligatorio", True) else "No",
            ]
            for ci, v in enumerate(vals, 1):
                cell_style(ws3.cell(row=row, column=ci, value=v), bg=bg)
            row += 1

    # Indicadores financieros de requisitos
    inds = financial.get("indicadores_financieros", []) or req_hab.get("financieros", [])
    for r in inds:
        vals = [
            "Financiero",
            r.get("nombre", "") or r.get("indicador", ""),
            r.get("formula", ""),
            r.get("valor_minimo", "") or r.get("valor_requerido", ""),
            "Sí",
        ]
        for ci, v in enumerate(vals, 1):
            cell_style(ws3.cell(row=row, column=ci, value=v), bg=C_BLUE_LIGHT)
        row += 1

    if row == 3:
        ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
        c = ws3.cell(row=3, column=1, value="No se identificaron requisitos habilitantes específicos")
        cell_style(c, bg=C_YELLOW)

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 4 — Análisis Financiero
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("4. Análisis Financiero")
    set_col_widths(ws4, [35, 35, 20, 25, 30])
    title_row(ws4, "💰 ANÁLISIS FINANCIERO", 1, 5, bg=C_GREEN_DARK)

    # Indicadores financieros
    hdr(ws4.cell(row=2, column=1, value="INDICADORES FINANCIEROS"), bg=C_GREEN_MED)
    for ci in range(2, 6):
        ws4.cell(row=2, column=ci).fill = PatternFill("solid", fgColor=C_GREEN_MED)

    hdrs4a = ["Indicador", "Fórmula", "Valor mínimo exigido", "Valor proponente", "Observaciones"]
    for ci, h in enumerate(hdrs4a, 1):
        hdr(ws4.cell(row=3, column=ci, value=h), bg=C_GREEN_MED, size=9)

    row = 4
    inds_fin = financial.get("indicadores_financieros", [])
    for ind in inds_fin:
        vals = [
            ind.get("nombre", "") or ind.get("indicador", ""),
            ind.get("formula", ""),
            ind.get("valor_minimo", "") or ind.get("valor_requerido", ""),
            "",    # Para que el proponente llene
            ind.get("observaciones", ""),
        ]
        bg = C_GREEN_LIGHT if row % 2 == 0 else C_WHITE
        for ci, v in enumerate(vals, 1):
            cell_style(ws4.cell(row=row, column=ci, value=v), bg=bg)
        row += 1

    # Capital de trabajo y patrimonio
    row += 1
    capital = financial.get("capital_trabajo", {})
    if capital:
        title_row(ws4, "CAPITAL DE TRABAJO", row, 5, bg=C_BLUE_MED)
        row += 1
        for k, v in capital.items():
            ws4.cell(row=row, column=1, value=k).font = Font(bold=True, size=9)
            ws4.cell(row=row, column=2, value=str(v))
            row += 1

    patrimonio = financial.get("patrimonio_minimo", "")
    if patrimonio:
        row += 1
        ws4.cell(row=row, column=1, value="Patrimonio mínimo exigido").font = Font(bold=True, size=9)
        ws4.cell(row=row, column=2, value=str(patrimonio))

    # Garantías
    row += 2
    garantias = legal.get("garantias", [])
    if garantias:
        title_row(ws4, "GARANTÍAS EXIGIDAS", row, 5, bg=C_ORANGE)
        row += 1
        hdrs_gar = ["Tipo de garantía", "Porcentaje / Valor", "Vigencia", "Observaciones", ""]
        for ci, h in enumerate(hdrs_gar, 1):
            hdr(ws4.cell(row=row, column=ci, value=h), bg=C_ORANGE)
        row += 1
        for g in garantias:
            vals = [
                g.get("tipo", "").title(),
                g.get("porcentaje", "") or g.get("valor", ""),
                g.get("vigencia", ""),
                g.get("observaciones", ""),
                "",
            ]
            for ci, v in enumerate(vals, 1):
                cell_style(ws4.cell(row=row, column=ci, value=v),
                           bg=C_YELLOW if row % 2 == 0 else C_WHITE)
            row += 1

    # Criterios de calificación
    criterios = legal.get("criterios_calificacion", [])
    if criterios:
        row += 1
        title_row(ws4, "CRITERIOS DE CALIFICACIÓN / EVALUACIÓN", row, 5, bg=C_BLUE_DARK)
        row += 1
        hdrs_crit = ["Criterio", "Tipo", "Puntaje máximo", "Descripción", ""]
        for ci, h in enumerate(hdrs_crit, 1):
            hdr(ws4.cell(row=row, column=ci, value=h), bg=C_BLUE_MED)
        row += 1
        total = 0
        for cr in criterios:
            pts = cr.get("puntaje_maximo", 0)
            total += pts if isinstance(pts, (int, float)) else 0
            vals = [cr.get("criterio",""), cr.get("tipo",""), pts, cr.get("descripcion",""), ""]
            for ci, v in enumerate(vals, 1):
                cell_style(ws4.cell(row=row, column=ci, value=v),
                           bg=C_BLUE_LIGHT if row % 2 == 0 else C_WHITE)
            row += 1
        # Fila total
        ws4.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=9)
        c_total = ws4.cell(row=row, column=3, value=total or legal.get("puntaje_total", 0))
        c_total.font = Font(bold=True, size=9)

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 5 — Productos y Especificaciones Técnicas
    # ══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("5. Productos y Especificaciones")
    set_col_widths(ws5, [8, 50, 12, 12, 15, 15, 20, 35])
    title_row(ws5, "🏷️ PRODUCTOS, CANTIDADES Y ESPECIFICACIONES TÉCNICAS", 1, 8, bg=C_BLUE_DARK)

    # Descripción general
    desc_gral = products.get("descripcion_general", "")
    if desc_gral:
        ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        c = ws5.cell(row=2, column=1, value=f"Descripción general: {desc_gral}")
        c.font = Font(italic=True, size=9)
        c.alignment = Alignment(wrap_text=True)
        ws5.row_dimensions[2].height = 30
        start_items = 4
    else:
        start_items = 3

    hdrs5 = ["#", "Descripción del ítem / Especificación técnica",
             "Cantidad", "Unidad", "Precio unitario est.", "Precio total est.",
             "Marca / Referencia", "Observaciones / Requisitos especiales"]
    for ci, h in enumerate(hdrs5, 1):
        hdr(ws5.cell(row=start_items - 1, column=ci, value=h), bg=C_BLUE_MED)

    items = products.get("items", [])
    row = start_items
    for i, item in enumerate(items, 1):
        bg = C_GREY_LIGHT if i % 2 == 0 else C_WHITE
        vals = [
            i,
            item.get("descripcion", "") or item.get("item", ""),
            item.get("cantidad", "") or item.get("quantity", ""),
            item.get("unidad", "") or item.get("unit", ""),
            item.get("precio_unitario", "") or item.get("unit_price", ""),
            item.get("precio_total", "") or item.get("total_price", ""),
            item.get("marca", "") or item.get("brand", "") or item.get("referencia", ""),
            item.get("observaciones", "") or item.get("especificaciones", "") or item.get("specs", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws5.cell(row=row, column=ci, value=v)
            cell_style(c, bg=bg, align="center" if ci in (1, 3, 4, 5, 6) else "left")
        ws5.row_dimensions[row].height = 20
        row += 1

    if not items:
        ws5.merge_cells(start_row=start_items, start_column=1, end_row=start_items, end_column=8)
        c = ws5.cell(row=start_items, column=1,
                     value="No se identificaron ítems de producto/servicio específicos en los documentos. "
                           "Revisar el estudio previo e invitación pública para el detalle de cantidades.")
        cell_style(c, bg=C_YELLOW)

    # Especificaciones técnicas adicionales
    specs = products.get("especificaciones_tecnicas", []) or products.get("technical_specs", [])
    if specs:
        row += 1
        title_row(ws5, "ESPECIFICACIONES TÉCNICAS ADICIONALES", row, 8, bg=C_BLUE_MED)
        row += 1
        for sp in specs:
            ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws5.cell(row=row, column=1,
                         value=f"• {sp}" if isinstance(sp, str) else
                               f"• {sp.get('especificacion', str(sp))}")
            c.alignment = Alignment(wrap_text=True)
            c.border = thin
            row += 1

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 6 — Riesgos e Indicadores
    # ══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("6. Riesgos e Indicadores")
    set_col_widths(ws6, [18, 50, 12, 12, 45])

    nivel = risks.get("nivel_riesgo_global", "N/A")
    score = risks.get("score_riesgo", 0)
    risk_bg = {"ALTO": C_RED_DARK, "MEDIO": C_ORANGE, "BAJO": C_GREEN_DARK}.get(nivel, C_BLUE_DARK)
    title_row(ws6, f"⚠️  ANÁLISIS DE RIESGOS — Nivel: {nivel} | Score: {score}/100", 1, 5, bg=risk_bg)

    row = 2
    # Indicadores de direccionamiento
    dirs = risks.get("indicadores_direccionamiento", [])
    if dirs:
        title_row(ws6, "🚩 INDICADORES DE POSIBLE DIRECCIONAMIENTO", row, 5, bg=C_RED_DARK)
        row += 1
        for ci, h in enumerate(["Tipo", "Descripción", "Severidad", "Probabilidad", "Recomendación"], 1):
            hdr(ws6.cell(row=row, column=ci, value=h), bg=C_RED_MED)
        row += 1
        for d in dirs:
            sev = d.get("severidad", "")
            sev_bg = {"ALTO": C_RED_LIGHT, "MEDIO": C_YELLOW, "BAJO": C_GREEN_LIGHT}.get(sev, C_WHITE)
            vals = [d.get("tipo",""), d.get("descripcion",""), sev,
                    d.get("probabilidad",""), d.get("recomendacion","")]
            for ci, v in enumerate(vals, 1):
                cell_style(ws6.cell(row=row, column=ci, value=v), bg=sev_bg)
            row += 1
        row += 1

    # Riesgos por categoría
    for cat_label, cat_key, cat_bg in [
        ("⚖️ RIESGOS JURÍDICOS",    "riesgos_juridicos",   C_BLUE_MED),
        ("💰 RIESGOS FINANCIEROS",  "riesgos_financieros", C_GREEN_MED),
        ("🔧 RIESGOS TÉCNICOS",     "riesgos_tecnicos",    C_ORANGE),
    ]:
        cat_risks = risks.get(cat_key, [])
        if not cat_risks:
            continue
        title_row(ws6, cat_label, row, 5, bg=cat_bg)
        row += 1
        for ci, h in enumerate(["Tipo", "Riesgo", "Probabilidad", "Impacto", "Mitigación"], 1):
            hdr(ws6.cell(row=row, column=ci, value=h), bg=cat_bg)
        row += 1
        for r in cat_risks:
            vals = [cat_label.split(" ", 1)[1][:15],
                    r.get("riesgo",""), r.get("probabilidad",""),
                    r.get("impacto",""), r.get("mitigacion","")]
            for ci, v in enumerate(vals, 1):
                cell_style(ws6.cell(row=row, column=ci, value=v),
                           bg=C_GREY_LIGHT if row % 2 == 0 else C_WHITE)
            row += 1
        row += 1

    # Alertas eliminatorias
    alertas = risks.get("alertas_eliminatorias", [])
    if alertas:
        title_row(ws6, "🚫 REQUISITOS ELIMINATORIOS", row, 5, bg=C_RED_DARK)
        row += 1
        for a in alertas:
            ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws6.cell(row=row, column=1, value=a.get("requisito","")).font = Font(bold=True, size=9)
            ws6.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            ws6.cell(row=row, column=3, value=a.get("descripcion",""))
            for ci in range(1, 6):
                ws6.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=C_RED_LIGHT)
                ws6.cell(row=row, column=ci).border = thin
            row += 1

    # Recomendaciones
    recs = risks.get("recomendaciones_oferta", [])
    if recs:
        row += 1
        title_row(ws6, "💡 RECOMENDACIONES PARA LA OFERTA", row, 5, bg=C_GREEN_DARK)
        row += 1
        for rec in recs:
            ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws6.cell(row=row, column=1, value=f"• {rec}")
            c.alignment = Alignment(wrap_text=True)
            c.fill  = PatternFill("solid", fgColor=C_GREEN_LIGHT)
            c.border = thin
            ws6.row_dimensions[row].height = 20
            row += 1

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 7 — Documentos analizados
    # ══════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("7. Documentos Analizados")
    set_col_widths(ws7, [8, 45, 20, 12, 12, 25])
    title_row(ws7, "📁 DOCUMENTOS DESCARGADOS Y ANALIZADOS", 1, 6, bg=C_BLUE_DARK)

    hdrs7 = ["#", "Nombre del archivo", "Tipo de documento", "Páginas", "Tamaño (KB)", "Fuente"]
    for ci, h in enumerate(hdrs7, 1):
        hdr(ws7.cell(row=2, column=ci, value=h), bg=C_BLUE_MED)

    if docs:
        for i, doc in enumerate(docs, 3):
            bg = C_GREY_LIGHT if i % 2 == 0 else C_WHITE
            vals = [i - 2,
                    doc.get("filename", ""),
                    doc.get("doc_type", ""),
                    doc.get("page_count", ""),
                    doc.get("size_kb", ""),
                    doc.get("source", "")]
            for ci, v in enumerate(vals, 1):
                cell_style(ws7.cell(row=i, column=ci, value=v), bg=bg)
    else:
        ws7.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
        c = ws7.cell(row=3, column=1,
                     value="No se pudieron descargar documentos. El análisis se basó en los metadatos del proceso SECOP II.")
        cell_style(c, bg=C_YELLOW)

    # ── Freeze panes y autofiltro ────────────────────────────────────────
    for ws in [ws3, ws5, ws6]:
        ws.freeze_panes = "A3"

    wb.save(str(output_path))


# ─────────────────────────────────────────────────────────────────────────────
# Excel básico (fallback si openpyxl falla en el completo)
# ─────────────────────────────────────────────────────────────────────────────

def _build_compliance_xlsx_basic(data: dict, output_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis SECOP II"

    process  = data.get("process_info", {})
    legal    = data.get("legal", {})
    risks    = data.get("risks", {})

    rows = [
        ("Referencia", process.get("referencia", "")),
        ("Entidad",    process.get("entidad", "")),
        ("Objeto",     legal.get("objeto_contrato", "")),
        ("Modalidad",  legal.get("modalidad_seleccion", "")),
        ("Valor",      legal.get("valor_estimado", "")),
        ("Riesgo",     risks.get("nivel_riesgo_global", "")),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)

    wb.save(str(output_path))


# ─────────────────────────────────────────────────────────────────────────────
# PDF unificado
# ─────────────────────────────────────────────────────────────────────────────

def _build_full_pdf(data: dict, output_path: Path):
    """Genera un PDF unificado. Intenta reportlab → fpdf2 → HTML con weasyprint."""
    try:
        _pdf_reportlab(data, output_path)
        return
    except ImportError:
        pass

    try:
        _pdf_fpdf2(data, output_path)
        return
    except ImportError:
        pass

    # Último recurso: generar HTML que se puede imprimir a PDF desde el navegador
    _pdf_html_fallback(data, output_path.with_suffix(".html"))
    raise ImportError("reportlab y fpdf2 no disponibles. Se generó HTML como alternativa.")


def _pdf_reportlab(data: dict, output_path: Path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable, PageBreak)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

    process  = data.get("process_info", {})
    legal    = data.get("legal", {})
    financial= data.get("financial", {})
    risks    = data.get("risks", {})
    products = data.get("products", {})
    docs     = data.get("documents", [])

    doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    normal.fontSize = 8
    normal.leading  = 12

    H1 = ParagraphStyle("H1", fontSize=14, fontName="Helvetica-Bold",
                        textColor=colors.HexColor("#1F4E79"), spaceAfter=8)
    H2 = ParagraphStyle("H2", fontSize=11, fontName="Helvetica-Bold",
                        textColor=colors.HexColor("#2E75B6"), spaceAfter=6, spaceBefore=10)
    H3 = ParagraphStyle("H3", fontSize=9, fontName="Helvetica-Bold",
                        textColor=colors.HexColor("#375623"), spaceAfter=4, spaceBefore=6)
    BODY = ParagraphStyle("BODY", fontSize=8, leading=12, spaceAfter=4)
    SMALL = ParagraphStyle("SMALL", fontSize=7, leading=10, textColor=colors.grey)

    BG_BLUE   = colors.HexColor("#BDD7EE")
    BG_GREEN  = colors.HexColor("#E2EFDA")
    BG_RED    = colors.HexColor("#FCE4D6")
    BG_HEADER = colors.HexColor("#1F4E79")
    FG_WHITE  = colors.white

    def tbl_style(header_rows=1, col_widths=None, alt_bg=None):
        s = [
            ("BACKGROUND", (0,0),  (-1, header_rows-1), BG_HEADER),
            ("TEXTCOLOR",  (0,0),  (-1, header_rows-1), FG_WHITE),
            ("FONTNAME",   (0,0),  (-1, header_rows-1), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0),  (-1,-1), 7),
            ("ALIGN",      (0,0),  (-1,-1), "LEFT"),
            ("VALIGN",     (0,0),  (-1,-1), "TOP"),
            ("GRID",       (0,0),  (-1,-1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, header_rows), (-1,-1),
             [alt_bg or colors.white, colors.HexColor("#F8F9FA")]),
        ]
        return TableStyle(s)

    story = []

    # ── Portada ──────────────────────────────────────────────────────────
    story.append(Paragraph("INFORME DE ANÁLISIS SECOP II", H1))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1F4E79")))
    story.append(Spacer(1, 0.3*cm))

    meta = [
        ["Referencia:", process.get("referencia", "N/A")],
        ["Entidad:",    process.get("entidad", "N/A")],
        ["Objeto:",     legal.get("objeto_contrato", "N/A")],
        ["Modalidad:",  legal.get("modalidad_seleccion", "N/A")],
        ["Valor estimado:", legal.get("valor_estimado", "N/A")],
        ["Duración:",   legal.get("duracion_contrato", "N/A")],
        ["Nivel de riesgo:", f"{risks.get('nivel_riesgo_global','N/A')} (Score: {risks.get('score_riesgo',0)}/100)"],
        ["Análisis generado:", datetime.now().strftime(DATE_FMT)],
    ]
    t = Table(meta, colWidths=[4*cm, 13*cm])
    t.setStyle(TableStyle([
        ("FONTNAME",   (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("GRID",       (0,0), (-1,-1), 0.25, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.HexColor("#EEF4FB"), colors.white]),
        ("VALIGN",     (0,0), (-1,-1), "TOP"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # ── Resumen ejecutivo riesgos ────────────────────────────────────────
    rs = risks.get("resumen_ejecutivo_riesgos", "")
    if rs:
        story.append(Paragraph("Resumen ejecutivo de riesgos", H2))
        story.append(Paragraph(rs, BODY))

    story.append(PageBreak())

    # ── Cronograma ───────────────────────────────────────────────────────
    cronograma = legal.get("cronograma", [])
    story.append(Paragraph("📅 Cronograma del Proceso", H2))
    if cronograma:
        cron_data = [["Evento / Actividad", "Fecha", "Observaciones"]]
        for ev in cronograma:
            cron_data.append([
                ev.get("evento", ""),
                ev.get("fecha", ""),
                ev.get("observaciones", ""),
            ])
        t = Table(cron_data, colWidths=[7*cm, 3*cm, 7*cm])
        t.setStyle(tbl_style())
        story.append(t)
    else:
        story.append(Paragraph("No se identificaron fechas en los documentos.", BODY))
    story.append(Spacer(1, 0.3*cm))

    # ── Requisitos habilitantes ──────────────────────────────────────────
    story.append(Paragraph("🔑 Requisitos Habilitantes", H2))
    req_hab = legal.get("requisitos_habilitantes", {})
    req_data = [["Tipo", "Requisito", "Cómo acreditar", "Oblig."]]
    for tipo, reqs in [("Jurídico", req_hab.get("juridicos",[])),
                       ("Técnico",  req_hab.get("tecnicos",[])),
                       ("Organiz.", req_hab.get("organizacionales",[]))]:
        for r in reqs:
            req_data.append([tipo, r.get("requisito",""),
                             r.get("documento_exigido","") or r.get("como_acreditar",""),
                             "Sí" if r.get("obligatorio", True) else "No"])

    inds = financial.get("indicadores_financieros", [])
    for r in inds:
        req_data.append(["Financiero",
                         r.get("nombre","") or r.get("indicador",""),
                         f"Fórmula: {r.get('formula','')} | Mín: {r.get('valor_minimo','')}",
                         "Sí"])

    if len(req_data) > 1:
        t = Table(req_data, colWidths=[2.5*cm, 6*cm, 6*cm, 1.5*cm])
        t.setStyle(tbl_style())
        story.append(t)
    else:
        story.append(Paragraph("No se identificaron requisitos habilitantes específicos.", BODY))
    story.append(Spacer(1, 0.3*cm))

    # ── Productos ────────────────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("🏷️ Productos, Cantidades y Especificaciones", H2))
    desc_gral = products.get("descripcion_general", "")
    if desc_gral:
        story.append(Paragraph(desc_gral, BODY))

    items = products.get("items", [])
    if items:
        prod_data = [["#", "Descripción", "Cantidad", "Unidad", "Precio U.", "Marca"]]
        for i, item in enumerate(items, 1):
            prod_data.append([
                str(i),
                item.get("descripcion", "") or item.get("item", ""),
                str(item.get("cantidad", "") or item.get("quantity", "")),
                item.get("unidad", "") or item.get("unit", ""),
                str(item.get("precio_unitario", "") or item.get("unit_price", "")),
                item.get("marca", "") or item.get("brand", "") or item.get("referencia", ""),
            ])
        t = Table(prod_data, colWidths=[0.8*cm, 7*cm, 1.5*cm, 1.5*cm, 2.5*cm, 3.7*cm])
        t.setStyle(tbl_style())
        story.append(t)
    else:
        story.append(Paragraph(
            "No se identificaron ítems específicos. Revisar el estudio previo e invitación pública.", BODY))

    # ── Riesgos ──────────────────────────────────────────────────────────
    story.append(PageBreak())
    nivel  = risks.get("nivel_riesgo_global", "N/A")
    score  = risks.get("score_riesgo", 0)
    story.append(Paragraph(f"⚠️ Análisis de Riesgos — Nivel: {nivel} | Score: {score}/100", H2))

    dirs = risks.get("indicadores_direccionamiento", [])
    if dirs:
        story.append(Paragraph("Indicadores de posible direccionamiento", H3))
        dir_data = [["Tipo", "Descripción", "Severidad", "Recomendación"]]
        for d in dirs:
            dir_data.append([d.get("tipo",""), d.get("descripcion",""),
                             d.get("severidad",""), d.get("recomendacion","")])
        t = Table(dir_data, colWidths=[3*cm, 6*cm, 2*cm, 6*cm])
        t.setStyle(tbl_style())
        story.append(t)
        story.append(Spacer(1, 0.2*cm))

    for cat_label, cat_key in [("Riesgos Jurídicos","riesgos_juridicos"),
                                ("Riesgos Financieros","riesgos_financieros"),
                                ("Riesgos Técnicos","riesgos_tecnicos")]:
        cat_risks = risks.get(cat_key, [])
        if not cat_risks:
            continue
        story.append(Paragraph(cat_label, H3))
        r_data = [["Riesgo", "Probabilidad", "Impacto", "Mitigación"]]
        for r in cat_risks:
            r_data.append([r.get("riesgo",""), r.get("probabilidad",""),
                           r.get("impacto",""), r.get("mitigacion","")])
        t = Table(r_data, colWidths=[5*cm, 2*cm, 2*cm, 8*cm])
        t.setStyle(tbl_style())
        story.append(t)
        story.append(Spacer(1, 0.2*cm))

    recs = risks.get("recomendaciones_oferta", [])
    if recs:
        story.append(Paragraph("💡 Recomendaciones para la Oferta", H3))
        for rec in recs:
            story.append(Paragraph(f"• {rec}", BODY))

    # ── Pie ──────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph(
        f"Análisis generado automáticamente por JIREHAI AI — {datetime.now().strftime(DATE_FMT)}. "
        "Este análisis es orientativo. Consulte siempre a un profesional especializado.",
        SMALL))

    doc.build(story)


def _pdf_fpdf2(data: dict, output_path: Path):
    from fpdf import FPDF

    process  = data.get("process_info", {})
    legal    = data.get("legal", {})
    risks    = data.get("risks", {})
    products = data.get("products", {})

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "INFORME ANALISIS SECOP II", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Referencia: {process.get('referencia','N/A')}", ln=True)
    pdf.cell(0, 6, f"Entidad: {process.get('entidad','N/A')}", ln=True)
    pdf.cell(0, 6, f"Valor estimado: {legal.get('valor_estimado','N/A')}", ln=True)
    pdf.cell(0, 6, f"Nivel de riesgo: {risks.get('nivel_riesgo_global','N/A')} "
                   f"(Score: {risks.get('score_riesgo',0)}/100)", ln=True)
    pdf.ln(4)

    # Items de productos
    items = products.get("items", [])
    if items:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 8, "PRODUCTOS Y CANTIDADES", ln=True)
        pdf.set_font("Helvetica", "", 8)
        for item in items:
            desc = str(item.get("descripcion","") or item.get("item",""))[:80]
            cant = str(item.get("cantidad","") or "")
            uni  = str(item.get("unidad","") or "")
            pdf.multi_cell(0, 5, f"- {desc} | Cant: {cant} {uni}")

    pdf.output(str(output_path))


def _pdf_html_fallback(data: dict, output_path: Path):
    """Genera un HTML bien formateado que se puede imprimir a PDF desde Chrome."""
    process  = data.get("process_info", {})
    legal    = data.get("legal", {})
    risks    = data.get("risks", {})
    products = data.get("products", {})
    financial= data.get("financial", {})

    items_html = ""
    for i, item in enumerate(products.get("items", []), 1):
        items_html += f"<tr><td>{i}</td><td>{item.get('descripcion','')}</td>" \
                      f"<td>{item.get('cantidad','')}</td><td>{item.get('unidad','')}</td>" \
                      f"<td>{item.get('precio_unitario','')}</td>" \
                      f"<td>{item.get('marca','')}</td></tr>"

    riesgos_html = ""
    for r in risks.get("riesgos_juridicos",[]) + risks.get("riesgos_financieros",[]) + risks.get("riesgos_tecnicos",[]):
        riesgos_html += f"<tr><td>{r.get('riesgo','')}</td><td>{r.get('probabilidad','')}</td>" \
                        f"<td>{r.get('impacto','')}</td><td>{r.get('mitigacion','')}</td></tr>"

    html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Análisis SECOP II - {process.get('referencia','')}</title>
<style>
  body {{font-family:Arial,sans-serif;font-size:11px;margin:20px}}
  h1 {{color:#1F4E79;font-size:16px}} h2 {{color:#2E75B6;font-size:13px;margin-top:20px}}
  table {{border-collapse:collapse;width:100%;margin-bottom:16px}}
  th {{background:#1F4E79;color:white;padding:5px;text-align:left;font-size:10px}}
  td {{border:1px solid #ccc;padding:4px;font-size:10px;vertical-align:top}}
  tr:nth-child(even) {{background:#f5f5f5}}
  .badge-alto {{background:#C00000;color:white;padding:2px 6px;border-radius:3px}}
  .badge-medio {{background:#ED7D31;color:white;padding:2px 6px;border-radius:3px}}
  .badge-bajo {{background:#375623;color:white;padding:2px 6px;border-radius:3px}}
  @media print {{ body{{margin:0}} }}
</style></head><body>
<h1>📋 INFORME ANÁLISIS SECOP II</h1>
<table>
  <tr><th>Campo</th><th>Valor</th></tr>
  <tr><td><b>Referencia</b></td><td>{process.get('referencia','N/A')}</td></tr>
  <tr><td><b>Entidad</b></td><td>{process.get('entidad','N/A')}</td></tr>
  <tr><td><b>Objeto</b></td><td>{legal.get('objeto_contrato','N/A')}</td></tr>
  <tr><td><b>Modalidad</b></td><td>{legal.get('modalidad_seleccion','N/A')}</td></tr>
  <tr><td><b>Valor estimado</b></td><td>{legal.get('valor_estimado','N/A')}</td></tr>
  <tr><td><b>Duración</b></td><td>{legal.get('duracion_contrato','N/A')}</td></tr>
  <tr><td><b>Nivel de riesgo</b></td>
    <td><span class="badge-{risks.get('nivel_riesgo_global','').lower()}">{risks.get('nivel_riesgo_global','N/A')}</span>
        Score: {risks.get('score_riesgo',0)}/100</td></tr>
</table>

<h2>🏷️ Productos y Cantidades</h2>
<table>
  <tr><th>#</th><th>Descripción</th><th>Cantidad</th><th>Unidad</th><th>Precio U.</th><th>Marca</th></tr>
  {items_html or '<tr><td colspan="6">No se identificaron ítems específicos</td></tr>'}
</table>

<h2>⚠️ Riesgos</h2>
<table>
  <tr><th>Riesgo</th><th>Probabilidad</th><th>Impacto</th><th>Mitigación</th></tr>
  {riesgos_html or '<tr><td colspan="4">No se identificaron riesgos específicos</td></tr>'}
</table>

<p style="color:#888;font-size:9px">
  Generado por JIREHAI AI — {datetime.now().strftime(DATE_FMT)}. Análisis orientativo.
</p>
</body></html>"""

    output_path.write_text(html, encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Constructores Markdown (sin cambios sustanciales)
# ─────────────────────────────────────────────────────────────────────────────

def _build_executive_summary(data: dict) -> str:
    process  = data.get("process_info", {})
    legal    = data.get("legal", {})
    financial= data.get("financial", {})
    risks    = data.get("risks", {})
    products = data.get("products", {})
    docs     = data.get("documents", [])
    now = datetime.now().strftime(DATE_FMT)

    lines = [
        "# 📋 Análisis SECOP II — Resumen Ejecutivo", "",
        f"> **Proceso:** {process.get('referencia','N/A')}  ",
        f"> **Entidad:** {process.get('entidad','N/A')}  ",
        f"> **Generado:** {now}  ",
        f"> **Documentos analizados:** {len(docs)}  ",
        "", "---", "",
        "## 🎯 Objeto del Contrato", "",
        legal.get("objeto_contrato", "_No identificado_"), "",
        f"**Alcance:** {legal.get('alcance', '_Ver documentos_')}", "",
        "---", "",
        "## 📊 Datos Generales", "",
        "| Campo | Valor |", "|-------|-------|",
        f"| Modalidad | {legal.get('modalidad_seleccion','N/A')} |",
        f"| Valor estimado | {legal.get('valor_estimado','N/A')} |",
        f"| Duración | {legal.get('duracion_contrato','N/A')} |",
        f"| Lugar | {legal.get('lugar_ejecucion','N/A')} |",
        f"| Tipo bien/servicio | {products.get('tipo_bien_servicio','N/A')} |",
        "", "---", "", "## 📦 Productos / Ítems", "",
    ]

    items = products.get("items", [])
    if items:
        lines += ["| # | Descripción | Cantidad | Unidad | Marca |",
                  "|---|-------------|---------|--------|-------|"]
        for item in items[:30]:
            lines.append(
                f"| {item.get('item','')} | {str(item.get('descripcion',''))[:60]} "
                f"| {item.get('cantidad','')} | {item.get('unidad','')} "
                f"| {item.get('marca','') or item.get('brand','')} |"
            )
        if len(items) > 30:
            lines.append(f"| ... | _{len(items)-30} ítems adicionales_ | | | |")
    else:
        lines.append("_No se identificaron ítems específicos_")

    nivel = risks.get("nivel_riesgo_global", "N/A")
    score = risks.get("score_riesgo", 0)
    em    = {"ALTO":"🔴","MEDIO":"🟡","BAJO":"🟢"}.get(nivel,"⚪")
    lines += [
        "", "---", "",
        f"## {em} Análisis de Riesgos", "",
        f"**Nivel:** {nivel} (Score: {score}/100)", "",
        risks.get("resumen_ejecutivo_riesgos",""), "",
        "", "---", "",
        f"*JIREHAI AI — {now}. Análisis orientativo.*",
    ]
    return "\n".join(lines)


def _build_legal_md(legal: dict, process_info: dict) -> str:
    now = datetime.now().strftime(DATE_FMT)
    ref = process_info.get("referencia", "N/A")
    lines = [
        f"# ⚖️ Análisis Jurídico — {ref}", f"_{now}_", "",
        "## Objeto y Alcance", "",
        f"**Objeto:** {legal.get('objeto_contrato','No identificado')}", "",
        f"**Alcance:** {legal.get('alcance','Ver documentos')}", "",
        "## Aspectos Generales", "",
        "| Aspecto | Detalle |", "|---------|---------|",
        f"| Modalidad | {legal.get('modalidad_seleccion','N/A')} |",
        f"| Justificación | {legal.get('justificacion_modalidad','N/A')} |",
        f"| Régimen | {legal.get('regimen_juridico','N/A')} |",
        f"| Duración | {legal.get('duracion_contrato','N/A')} |",
        f"| Valor | {legal.get('valor_estimado','N/A')} |",
        f"| Lugar | {legal.get('lugar_ejecucion','N/A')} |",
    ]

    criterios = legal.get("criterios_calificacion", [])
    if criterios:
        lines += ["", "## Criterios de Calificación", "",
                  "| Criterio | Puntaje | Tipo |", "|----------|---------|------|"]
        for c in criterios:
            lines.append(f"| {c.get('criterio','')} | {c.get('puntaje_maximo',0)} | {c.get('tipo','')} |")
        lines.append(f"| **TOTAL** | **{legal.get('puntaje_total',0)}** | |")

    garantias = legal.get("garantias", [])
    if garantias:
        lines += ["", "## Garantías", ""]
        for g in garantias:
            lines.append(f"- **{g.get('tipo','').title()}:** {g.get('porcentaje','')} — {g.get('vigencia','')}")

    exp = legal.get("experiencia_exigida", {})
    if exp.get("tipo_experiencia"):
        lines += ["", "## Experiencia Requerida", "",
                  f"- Años mínimos: {exp.get('años_minimos',0)}",
                  f"- Tipo: {exp.get('tipo_experiencia','')}",
                  f"- Acreditación: {exp.get('como_acreditar','')}"]

    return "\n".join(lines)


def _build_risk_report_md(risks: dict, process_info: dict) -> str:
    now   = datetime.now().strftime(DATE_FMT)
    ref   = process_info.get("referencia", "N/A")
    level = risks.get("nivel_riesgo_global", "N/A")
    score = risks.get("score_riesgo", 0)
    em    = {"ALTO":"🔴","MEDIO":"🟡","BAJO":"🟢"}.get(level,"⚪")

    lines = [
        f"# {em} Reporte de Riesgos — {ref}", f"_{now}_", "",
        f"**Nivel:** {level} | **Score:** {score}/100", "",
        "## Resumen", "",
        risks.get("resumen_ejecutivo_riesgos","_Sin resumen_"), "",
    ]

    for title, key in [("⚠️ Indicadores de Direccionamiento","indicadores_direccionamiento"),
                       ("⚖️ Riesgos Jurídicos","riesgos_juridicos"),
                       ("💰 Riesgos Financieros","riesgos_financieros"),
                       ("🔧 Riesgos Técnicos","riesgos_tecnicos")]:
        items = risks.get(key, [])
        if not items:
            continue
        lines += [f"## {title}", ""]
        for item in items:
            sev = item.get("severidad","") or item.get("probabilidad","")
            s_em = {"ALTO":"🔴","MEDIO":"🟡","BAJO":"🟢"}.get(str(sev).upper(),"")
            desc = item.get("descripcion","") or item.get("riesgo","")
            rec  = item.get("recomendacion","") or item.get("mitigacion","")
            lines.append(f"### {s_em} {item.get('tipo','') or desc[:60]}")
            if desc and desc != item.get('tipo',''):
                lines.append(desc)
            if rec:
                lines.append(f"**Mitigación:** {rec}")
            lines.append("")

    recs = risks.get("recomendaciones_oferta", [])
    if recs:
        lines += ["## 💡 Recomendaciones para la Oferta", ""]
        for r in recs:
            lines.append(f"- {r}")

    return "\n".join(lines)


def _write_json(data: dict, path: Path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
